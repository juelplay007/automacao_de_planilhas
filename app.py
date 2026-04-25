from openpyxl import load_workbook, Workbook

planilha_vendas = load_workbook('vendas_de_lanches.xlsx')
pagina_vendas = planilha_vendas ['Sheet']

for linha in pagina_vendas.iter_rows(values_only=True):
    print(linha)

    planilha_contas = Workbook() 
    pagina1 = planilha_contas.active

with open('anotacao.txt', 'r', encoding='utf-8') as arquivo:
        for linha in arquivo:
            pagina1.append(linha.split(','))

            planilha_contas.save('anotacao2.xlsx')